import os
import xlwt
import difflib
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import xlrd
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

#多张表的统计
def classify(path):
    file = os.listdir(path)#遍历文件
    list3=[]
    list4=[]
    for i in file:
        name = os.path.join(path, i)
        a=os.path.isdir(name) #盘算是否是文件夹
        if a==True:  #如果是文件夹，继续运行
            for filename in os.listdir(name):
                if filename == '总结果.xls':
                    # 读取Excel文件
                    df = pd.read_excel(os.path.join(name, filename),sheet_name= '分类')
                    list1=df['类'].tolist()
                    list2= df['价格'].tolist()
                    for j in range(len(list1)):
                        list3.append(list1[j])
                        list4.append(list2[j])

    workbook = xlwt.Workbook(encoding='ufo-8')
    worksheet = workbook.add_sheet('完成')
    i=0#将名字一样的价格加在一起
    while i < len(list3) - 1:  #内容
        j = i+1
        while j<=len(list4) - 1: #价格
            a = difflib.SequenceMatcher(None, list3[i], list3[j]).quick_ratio()  # 判断相似度
            if a>0.65:
                list3.pop(j)
                list4[i]=list4[i]+list4[j]
                list4.pop(j)
                j=j-1
            j=j+1
        i=i+1
    worksheet.write(0,0,'类')
    worksheet.write(0, 1, '价格')
    for i in range(len(list3)):
        worksheet.write(i+1,0,list3[i])
        worksheet.write(i+1, 1, list4[i])
    workbook.save(path+'/统计.xls')


    df = pd.read_excel(path + '/统计.xls')  # 如果没有标题行，设置为header=None
    list1=df['类'].tolist()
    list2=df['价格'].tolist()
    for i in range(len(list2)):
        list2[i]=int(float(list2[i]))
    y = np.array(list2)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    plt.figure(figsize=(3, 2))  # 宽度为8英寸，高度为6英寸
    plt.pie(y,
            labels=list1,  # 设置饼图标签
            autopct='%.2f%%',  # 格式化输出百分比
            )
    plt.title("统计结果")#写标题
    plt.savefig(path+'/tu.png',dpi=100)

    workbook_xls = xlrd.open_workbook(path + '/统计.xls')
    sheet_xls = workbook_xls.sheet_by_index(0)
    # 创建一个新的xlsx文件
    workbook_xlsx = openpyxl.Workbook()
    sheet_xlsx = workbook_xlsx.active
    # 复制数据
    for row in range(0, sheet_xls.nrows):
        for col in range(0, sheet_xls.ncols):
            sheet_xlsx.cell(row=row + 1, column=col + 1, value=sheet_xls.cell_value(row, col))
    # 插入图片
    img = PILImage.open(path+'/tu.png')
    img_path = path+'/tu.png'
    img.save(img_path)
    img = Image(img_path)
    sheet_xlsx.add_image(img, 'D1')
    # 保存xlsx文件
    workbook_xlsx.save(path + '/统计.xlsx')
    os.remove(path + '/统计.xls')
    return img_path
#
if __name__ == "__main__":
    path=r"D:\2\xgk\result"

    classify(path)
